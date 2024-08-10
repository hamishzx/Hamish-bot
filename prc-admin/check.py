# ！/usr/bin/env python
# -*-coding:Utf-8 -*-
# Time: 7 Aug 2024 12:51
# Name: check.py
# Author: CHAU SHING SHING HAMISH

import re

import openpyxl
import pywikibot
from tqdm import tqdm

site = pywikibot.Site('zh', 'wikipedia')
site.login()


def construct_table(data, t):
    table = ''
    if data[2] != '':
        data[2] = '[[' + data[2] + '|list]] - [[' + data[3] + '|data]]'
    if t == 'shared':
        table += '\n|-\n| {0} || {1} || 1 || 1 || {2} || {3}'.format(data[0], data[1], data[2], data[4])
    elif t == 'data':
        table += '\n|-\n| {0} || {1} || 1 || 0 || {2} || {3}'.format(data[0], data[1], data[2], data[4])
    elif t == 'page':
        table += '\n|-\n| {0} || {1} || 0 || 1 || {2} || {3}'.format(data[0], data[1], data[2], data[4])
    return table


def show_subdivisions(*args):
    sheet = args[0]
    process_admin_name = args[1]
    subdivisions = []

    if len(args) == 2:
        # case show city
        for r in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            if r[4:7] == ('00', '000', '000'):
                subdivisions.append(r)
    elif len(args) == 3:
        # case show county
        city_to_show = args[2]
        for r in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            if r[5:7] == ('000', '000') and r[3] == city_to_show:
                subdivisions.append(r)
    elif len(args) == 4:
        # case show town
        city_to_show = args[2]
        county_to_show = args[3]
        for r in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            if r[6] == '000' and r[3:5] == (city_to_show, county_to_show):
                subdivisions.append(r)
    elif len(args) == 5:
        # case show village
        city_to_show = args[2]
        county_to_show = args[3]
        town_to_show = args[4]
        for r in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            if r[6] != '000' and r[3:6] == (city_to_show, county_to_show, town_to_show):
                subdivisions.append(r)
    return subdivisions


def compare_data(name, data, subs, level):
    sub_code_sheet = [r[level] for r in subs]
    data_process = set(sub_code_sheet)
    subs_process = set(data)
    shared = data_process & subs_process
    data_process = [x for x in data_process if x not in shared]
    subs_process = [x for x in subs_process if x not in shared]

    result = [shared, data_process, subs_process]
    return result


def verify_integrity():
    base_page = pywikibot.Page(site, 'Template:PRC_admin/list/00/00/00/000/000')
    pattern = re.compile(r'\|(\d{2,3})')
    provinces = pattern.findall(base_page.text)

    for province in tqdm(provinces, desc='Processing Provinces'):
        table = '{| class="wikitable sortable"\n|-\n! 行政區劃代碼 !! 行政區劃名稱 !! 站內 !! 數據庫 !! 連結 !! 備註'
        province_data = openpyxl.load_workbook('Administration_' + province + '.xlsx')
        province_data_sheet = province_data.active
        province_list_page = pywikibot.Page(site, 'Template:PRC_admin/list/' + province + '/00/00/000/000')
        province_data_page = pywikibot.Page(site, 'Template:PRC_admin/data/' + province + '/00/00/000/000')
        province_name = re.findall(r'name=(.*)\|', province_data_page.text)[0]
        cities = pattern.findall(province_list_page.text)
        province_subdivisions = show_subdivisions(province_data_sheet, province_name)
        province_compare_result = compare_data(province_name, cities, province_subdivisions, 3)
        for c in province_compare_result[2]:
            c_data = [p for p in province_subdivisions if p[3] == c]
            table += construct_table([c_data[0][0], c_data[0][1], '', '', '模板頁缺失'], 'page')

        for city in tqdm(cities, desc='Processing Cities'):
            city_data = [p for p in province_subdivisions if p[3] == city]
            city_list_page = pywikibot.Page(site, 'Template:PRC_admin/list/' + province + '/' + city + '/00/000/000')
            city_data_page = pywikibot.Page(site, 'Template:PRC_admin/data/' + province + '/' + city + '/00/000/000')
            try:
                city_name = re.findall(r'name=(.*)\|', city_data_page.text)[0]
            except IndexError:
                city_name = 'IndexError'

            if city in province_compare_result[0]:
                if city_name == city_data[0][1]:
                    if city_name == 'IndexError':
                        table += construct_table([city_data[0][0], city_name, city_list_page.title(),
                                                  city_data_page.title(), 'IndexError，檢查數據準確性'], 'shared')
                    else:
                        table += construct_table([city_data[0][0], city_name, city_list_page.title(),
                                                city_data_page.title(), ''], 'shared')
                else:
                    table += construct_table([city_data[0][0], city_name, city_list_page.title(),
                                              city_data_page.title(), '名稱或已變更'], 'shared')
            elif city in province_compare_result[1]:
                table += construct_table([city_data[0][0], city_name, city_list_page, city_data_page, '數據庫缺失']
                                         , 'data')

            counties = pattern.findall(city_list_page.text)
            city_subdivisions = show_subdivisions(province_data_sheet, city_name, city)
            city_compare_result = compare_data(city_name, counties, city_subdivisions, 4)

            for c in city_compare_result[2]:
                c_data = [p for p in city_subdivisions if p[4] == c]
                table += construct_table([c_data[0][0], c_data[0][1], '', '', '模板頁缺失'], 'page')

            for county in tqdm(counties, desc='Processing Counties'):
                county_data = [p for p in city_subdivisions if p[4] == county]
                county_list_page = pywikibot.Page(site,
                                                  'Template:PRC_admin/list/' + province + '/' + city + '/' + county + '/000/000')
                county_data_page = pywikibot.Page(site,
                                                  'Template:PRC_admin/data/' + province + '/' + city + '/' + county + '/000/000')
                try:
                    county_name = re.findall(r'name=(.*)\|', county_data_page.text)[0]
                except IndexError:
                    county_name = 'IndexError'

                if county in city_compare_result[0]:
                    if county_name == county_data[0][1]:
                        table += construct_table([county_data[0][0], county_name, county_list_page.title(),
                                                  county_data_page.title(), ''], 'shared')
                    else:
                        table += construct_table([county_data[0][0], county_name, county_list_page.title(),
                                                  county_data_page.title(), '名稱或已變更'], 'shared')
                elif county in city_compare_result[1]:
                    table += construct_table([county_data[0][0], county_name, county_list_page.title(),
                                              county_data_page.title(), '數據庫缺失'], 'data')

                towns = pattern.findall(county_list_page.text)
                county_subdivisions = show_subdivisions(province_data_sheet, county_name, city, county)
                county_compare_result = compare_data(county_name, towns, county_subdivisions, 5)

                for c in county_compare_result[2]:
                    c_data = [p for p in county_subdivisions if p[5] == c]
                    table += construct_table([c_data[0][0], c_data[0][1], '', '', '模板頁缺失'], 'page')

                for town in tqdm(towns, desc='Processing Towns'):
                    town_data = [p for p in county_subdivisions if p[5] == town]
                    town_list_page = pywikibot.Page(site,
                                                    'Template:PRC_admin/list/' + province + '/' + city + '/' + county + '/' + town + '/000')
                    town_data_page = pywikibot.Page(site,
                                                    'Template:PRC_admin/data/' + province + '/' + city + '/' + county + '/' + town + '/000')
                    try:
                        town_name = re.findall(r'name=(.*)\|', town_data_page.text)[0]
                    except IndexError:
                        town_name = 'IndexError'

                    if town in county_compare_result[0]:
                        if town_name == town_data[0][1]:
                            table += construct_table([town_data[0][0], town_name, town_list_page.title(),
                                                      town_data_page.title(), ''], 'shared')
                        else:
                            table += construct_table([town_data[0][0], town_name, town_list_page.title(),
                                                      town_data_page.title(), '名稱或已變更'], 'shared')
                    elif town in county_compare_result[1]:
                        table += construct_table([town_data[0][0], town_name, town_list_page.title(),
                                                  town_data_page.title(), '數據庫缺失'], 'data')

                    villages = pattern.findall(town_list_page.text)
                    town_subdivisions = show_subdivisions(province_data_sheet, town_name, city, county, town)
                    town_compare_result = compare_data(town_name, villages, town_subdivisions, 6)

                    for c in town_compare_result[2]:
                        c_data = [p for p in town_subdivisions if p[6] == c]
                        table += construct_table([c_data[0][0], c_data[0][1], '', '', '模板頁缺失'], 'page')
        table += '\n|}'
        user_page = pywikibot.Page(site, 'User:Hamish-bot/PRC_admin/' + province)
        user_page.text = table
        user_page.save(summary='Bot: Update PRC admin list result')


if __name__ == '__main__':
    verify_integrity()
